from csv import reader as csv_reader
from re import sub
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit


def custom_quit(msg: str) -> None:
    print(msg)
    quit()




class UserInterface:
    file_name: str
    profession_name: str

    def __init__(self, file_name: str = None):
        if file_name is not None:
            self.file_name = file_name
        else:
            self.file_name = "../vacancies.csv"
        self.profession_name = 'Программист'

class Translator:
    AZN: str = "Манаты"
    BYR: str = "Белорусские рубли"
    EUR: str = "Евро"
    GEL: str = "Грузинский лари"
    KGS: str = "Киргизский сом"
    KZT: str = "Тенге"
    RUR: str = "Рубли"
    UAH: str = "Гривны"
    USD: str = "Доллары"
    UZS: str = "Узбекский сум"
    currency_to_rub: {str, float} = {
        "Манаты": 35.68,
        "Белорусские рубли": 23.91,
        "Евро": 59.90,
        "Грузинский лари": 21.74,
        "Киргизский сом": 0.76,
        "Тенге": 0.13,
        "Рубли": 1,
        "Гривны": 1.64,
        "Доллары": 60.66,
        "Узбекский сум": 0.0055,
    }

    def translate(self, key: str, dict_name: str = None) -> str:
        if dict_name is not None:
            return self.__getattribute__(dict_name)[key]
        return self.__getattribute__(key)

    def translate_currency_to_rub(self, currency: str) -> int or float:
        return self.currency_to_rub[currency]


class CSV:
    data: csv_reader
    title: list
    rows: list

    def __init__(self, file_name: str):
        with open(file_name, 'r', newline='', encoding='utf-8-sig') as file:
            self.data = csv_reader(file)
            try:
                self.title = next(self.data)
            except StopIteration:
                custom_quit('Пустой файл')

            self.rows = [row for row in self.data
                         if len(list(filter(lambda word: word != '', row))) == len(self.title)]

            if len(self.rows) == 0:
                custom_quit('Нет данных')


class Salary:
    salary_from: int
    salary_to: int
    salary_currency: str

    def set_field(self, key: str, value: str):
        if key == 'salary_currency':
            value = translator.translate(value)
        if key in ['salary_from', 'salary_to']:
            value = float(value)
        self.__setattr__(key, value)

    def get_average_in_rur(self) -> int:
        return int(translator.translate_currency_to_rub(self.salary_currency) *
                   (float(self.salary_from) + float(self.salary_to)) // 2)


class Vacancy:
    name: str
    salary: Salary
    area_name: str
    published_at: int

    @staticmethod
    def get_correct_field(key: str, value: str or list) -> int or str:
        if key == 'published_at':
            big, small = value[:19].split('T')
            year, month, day = big.split('-')
            # hours, minutes, seconds = small.split(':')
            return int(year)
        else:
            return value

    def __init__(self, fields: dict):
        for key, value in fields.items():
            if not self.check_salary(key, value):
                self.__setattr__(key, self.get_correct_field(key, value))


    def check_salary(self, key: str, value: str) -> bool:
        is_salary = False
        if key in ['salary_from', 'salary_to', 'salary_currency']:
            if not hasattr(self, 'salary'):
                self.salary = Salary()
            self.salary.set_field(key, value)
            is_salary = True
        return is_salary

    def get_field(self, field: str):
        if field in 'salary':
            return self.salary.get_average_in_rur()
        return self.__getattribute__(field)






class DataSet:
    profession_name: str
    profession_count: int
    vacancies: List[Vacancy]
    salary_by_years: {int, list}
    vacancies_by_years: {int, int}
    profession_salary_by_years: {int, list}
    profession_vacancies_by_years: {int, int}
    salaries_by_cities: {str, list}
    ratio_vacancy_by_cities: {str, float}
    city_vacancies_count: {str, int}

    def __init__(self, vacs: list, prof_name: str):
        self.profession_name = prof_name
        self.profession_count = 0
        self.vacancies = vacs
        self.salary_by_years = {}
        self.vacancies_by_years = {}
        self.profession_salary_by_years = {}
        self.profession_vacancies_by_years = {}
        self.salaries_by_cities = {}
        self.ratio_vacancy_by_cities = {}
        self.city_vacancies_count = {}

        self._get_data()

    def _get_data(self) -> None:
        for vac in self.vacancies:
            self.process_vacancies_count('city_vacancies_count', 'area_name', vac)
        for vac in self.vacancies:
            self.process_salary('salary_by_years', 'published_at', vac)
            self.process_vacancies_count('vacancies_by_years', 'published_at', vac)
            if self.profession_name in vac.name:
                self.profession_count += 1
                self.process_salary('profession_salary_by_years', 'published_at', vac)
                self.process_vacancies_count('profession_vacancies_by_years', 'published_at', vac)
            self.process_salary('salaries_by_cities', 'area_name', vac)
            self.process_vacancies_count('ratio_vacancy_by_cities', 'area_name', vac)

        self.set_correct_cities_data()

    def process_salary(self, dict_name: str, field: str, vac: Vacancy) -> None:
        d = self.__getattribute__(dict_name)
        f = vac.get_field(field)
        if f not in d.keys():
            d[f] = [vac.salary.get_average_in_rur(), 1]
        else:
            d[f][0] += vac.salary.get_average_in_rur()
            d[f][1] += 1

    def process_vacancies_count(self, dict_name: str, field: str, vac: Vacancy) -> None:
        d = self.__getattribute__(dict_name)
        f = vac.get_field(field)
        if f not in d.keys():
            d[f] = 1
        else:
            d[f] += 1

    def set_correct_cities_data(self):
        for key, value in self.ratio_vacancy_by_cities.items():
            self.ratio_vacancy_by_cities[key] = round(value / len(self.vacancies), 4)

        d1 = dict(sorted(self.salaries_by_cities.items(), key=lambda i: i[1][1] / i[1][0]))
        self.salaries_by_cities = self.get_first_ten_correct(d1)

        d2 = dict(sorted(self.ratio_vacancy_by_cities.items(), key=lambda i: i[1], reverse=True))
        self.ratio_vacancy_by_cities = self.get_first_ten_correct(d2)

    def get_first_ten_correct(self, d: dict) -> dict:
        count = 0
        res = {}
        for key, value in d.items():
            if count == 10:
                break
            if self.city_vacancies_count[key] >= len(self.vacancies) // 100:
                res[key] = value
                count += 1
        return res

    def get_data(self) -> dict:
        salaries_by_years, vacancies_by_years = [], []
        salaries_by_cities, ratio_vacancies_by_cities = {}, {}
        to_print: {str, dict} \
            = {"Уровень зарплат по годам": self.salary_by_years,
               "Количество вакансий по годам": self.vacancies_by_years,
               "Уровень зарплат по годам для выбранной профессии": self.profession_salary_by_years,
               "Количество вакансий по годам для выбранной профессии": self.profession_vacancies_by_years,
               "Уровень зарплат по городам": self.salaries_by_cities,
               "Доля вакансий по городам": self.ratio_vacancy_by_cities}
        for key, value in to_print.items():
            if len(value) == 0:
                value = {k: 0 for k in self.salary_by_years.keys()}
            for k, v in value.items():
                if type(v) is list:
                    value[k] = v[0] // v[1]
            if 'Уровень зарплат по годам' in key:
                salaries_by_years.append(value)
            elif 'Количество вакансий по годам' in key:
                vacancies_by_years.append(value)
            elif 'Уровень зарплат по городам' in key:
                salaries_by_cities = value
            else:
                ratio_vacancies_by_cities = value

        return {"Уровень зарплат по годам": salaries_by_years,
                "Количество вакансий по годам": vacancies_by_years,
                "Уровень зарплат по городам": salaries_by_cities,
                "Доля вакансий по городам": ratio_vacancies_by_cities}


class Report:
    workbook: Workbook
    data: dict

    def __init__(self, data: dict, **kwargs):
        self.workbook = Workbook()
        self.data = data
        for key, value in kwargs.items():
            self.__setattr__(key, value)


    def fill_with_statistics(self) -> None:
        self.fill_salaries_statistics()
        self.fill_cities_statistics()

    # region Excel
    def generate_excel(self, file_name: str) -> None:
        self.fill_with_statistics()
        self.workbook.save(file_name)



    def fill_salaries_statistics(self) -> None:
        ws = self.workbook.active
        ws.title = 'Статистика по годам'
        salaries_by_years = self.data["Уровень зарплат по годам"][0]
        vacancies_by_years = self.data["Количество вакансий по годам"][0]
        profession_salaries_by_years = self.data["Уровень зарплат по годам"][1]
        profession_vacancies_by_years = self.data["Количество вакансий по годам"][1]

        self.fill_column('Год', list(salaries_by_years.keys()),
                         [cell[0] for cell in ws['A1':f'A{len(salaries_by_years) + 1}']])

        self.fill_column('Средняя зарплата', list(salaries_by_years.values()),
                         [cell[0] for cell in ws['B1':f'B{len(salaries_by_years) + 1}']])
        self.fill_column(f'Средняя зарплата - {ds.profession_name}', list(profession_salaries_by_years.values()),
                         [cell[0] for cell in ws['C1':f'C{len(profession_salaries_by_years) + 1}']])

        self.fill_column('Количество вакансий', list(vacancies_by_years.values()),
                         [cell[0] for cell in ws['D1':f'D{len(vacancies_by_years) + 1}']])
        self.fill_column(f'Количество вакансий - {ds.profession_name}', list(profession_vacancies_by_years.values()),
                         [cell[0] for cell in ws['E1':f'E{len(profession_vacancies_by_years) + 1}']])

        self.update_worksheet_settings(ws)

    def fill_cities_statistics(self) -> None:
        self.workbook.create_sheet("Статистика по городам")
        ws = self.workbook["Статистика по городам"]
        salaries_by_cities = self.data["Уровень зарплат по городам"]
        vacs_ratio_by_cities = self.data["Доля вакансий по городам"]

        self.fill_column('Город', list(salaries_by_cities.keys()),
                         [cell[0] for cell in ws['A1':f'A{len(salaries_by_cities) + 1}']])
        self.fill_column('Уровень зарплат', list(salaries_by_cities.values()),
                         [cell[0] for cell in ws['B1': f'B{len(salaries_by_cities) + 1}']])

        self.fill_column('Город', list(vacs_ratio_by_cities.keys()),
                         [cell[0] for cell in ws['D1':f'D{len(vacs_ratio_by_cities) + 1}']])
        self.fill_column('Доля вакансий', list(vacs_ratio_by_cities.values()),
                         [cell[0] for cell in ws['E1': f'E{len(vacs_ratio_by_cities) + 1}']])

        self.set_column_percent([cell[0] for cell in ws['E2': f'E{len(vacs_ratio_by_cities) + 1}']])
        self.update_worksheet_settings(ws)

    @staticmethod
    def set_column_percent(column: list) -> None:
        for cell in column:
            cell.number_format = FORMAT_PERCENTAGE_00

    def update_worksheet_settings(self, ws) -> None:
        self.set_borders(ws)
        self.set_column_width(ws)

    @staticmethod
    def fill_column(header: str, data: list, column_cells: list) -> None:
        column_cells[0].value = header
        for cell, value in zip(column_cells[1:], data):
            cell.value = value



    @staticmethod
    def set_borders(ws) -> None:
        isFirstRow = True
        for row in ws.rows:
            for cell in row:
                if not cell.value:
                    continue
                cell.border = Border(top=Side(border_style="thin", color="000000"),
                                     left=Side(border_style="thin", color="000000"),
                                     right=Side(border_style="thin", color="000000"),
                                     bottom=Side(border_style="thin", color="000000"))
                if isFirstRow:
                    cell.font = Font(bold=True)
            isFirstRow = False

    @staticmethod
    def set_column_width(ws) -> None:
        a = {0: "A", 1: "B", 2: "C", 3: "D", 4: "E", 6: "F", 7: "G"}
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value)) + 1))

        for col, value in dims.items():
            ws.column_dimensions[a[col - 1]].width = value

    # endregion
    # region Plot

    def generate_image(self, file_name: str, show_result: bool = False) -> None:
        self.draw_graphs()
        plt.tight_layout()
        plt.savefig(file_name, dpi=300)
        if show_result:
            plt.show()

    def draw_graphs(self) -> None:
        figure, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2)
        self.draw_bar_graph(ax1, "Уровень зарплат по годам")
        self.draw_bar_graph(ax2, "Количество вакансий по годам")
        self.draw_invert_bar_graph(ax3, "Уровень зарплат по городам")
        self.draw_pie_graph(ax4, "Доля вакансий по городам")

    def draw_bar_graph(self, subplot, name: str) -> None:
        bar_width = 0.4
        first_label = 'средняя з/п'
        second_label = f'з/п {ui.profession_name}'
        if name == "Количество вакансий по годам":
            first_label = "Количество вакансий"
            second_label = f"Количество вакансий\n{ui.profession_name}"

        average_by_years: dict = self.data[name][0]
        profession_average_by_years: dict = self.data[name][1]

        X_axis = np.arange(len(average_by_years.keys()))

        subplot.bar(X_axis - bar_width / 2, average_by_years.values(), width=bar_width, label=first_label)
        subplot.bar(X_axis + bar_width / 2, profession_average_by_years.values(),
                    width=bar_width, label=second_label)
        subplot.set_xticks(X_axis, average_by_years.keys())
        subplot.set_xticklabels(average_by_years.keys(), rotation='vertical', va='top', ha='center')

        subplot.set_title(name)
        subplot.grid(True, axis='y')
        subplot.tick_params(axis='both', labelsize=8)
        subplot.legend(fontsize=8)

    def draw_invert_bar_graph(self, subplot, name: str) -> None:
        subplot.invert_yaxis()
        courses = list(self.data[name].keys())
        courses = [label.replace(' ', '\n').replace('-', '-\n') for label in courses]
        values = list(self.data[name].values())
        subplot.barh(courses, values)
        subplot.set_yticklabels(courses, fontsize=6, va='center', ha='right')

        subplot.set_title(name)
        subplot.grid(True, axis='x')
        subplot.tick_params(axis='both', labelsize=8)

    def draw_pie_graph(self, subplot, name: str) -> None:
        data = self.data[name]
        other = 1 - sum((list(data.values())))
        new_dic = {'Другие': other}
        new_dic.update(data)

        labels = list(new_dic.keys())
        sizes = list(new_dic.values())

        subplot.set_title(name)
        subplot.pie(sizes, labels=labels, textprops={'fontsize': 6})
        subplot.axis('scaled')

    # endregion
    # region PDF

    def generate_pdf(self, name: str):
        image_file = "graph.png"
        header_year = ["Год", "Средняя зарплата", f"Средняя зарплата - {ds.profession_name}", "Количество вакансий",
                       f"Количество вакансий - {ds.profession_name}"]
        header_city = ["Город", "Уровень зарплат", '', "Город", "Доля вакансий"]

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")

        salaries_by_years = self.data["Уровень зарплат по годам"][0]
        vacancies_by_years = self.data["Количество вакансий по годам"][0]
        profession_salaries_by_years = self.data["Уровень зарплат по годам"][1]
        profession_vacancies_by_years = self.data["Количество вакансий по годам"][1]
        salaries_by_cities = self.data["Уровень зарплат по городам"]
        ratio_vacancy_by_cities = {city: str(f'{ratio * 100:,.2f}%').replace('.', ',')
                                   for city, ratio in self.data["Доля вакансий по городам"].items()}

        salary_data = {year: [salary, count, salary_vac, count_vac]
                       for year, salary, count, salary_vac, count_vac in zip(salaries_by_years.keys(),
                                                                             salaries_by_years.values(),
                                                                             vacancies_by_years.values(),
                                                                             profession_salaries_by_years.values(),
                                                                             profession_vacancies_by_years.values())}
        city_data = {index: [salary_city, salary, ratio_city, ratio]
                     for index, (salary_city, salary, ratio_city, ratio) in
                     enumerate(zip(salaries_by_cities.keys(),
                                   salaries_by_cities.values(),
                                   ratio_vacancy_by_cities.keys(),
                                   ratio_vacancy_by_cities.values()))}

        pdf_template = template.render(
            {'image_file': image_file,
             'image_style': 'style="max-width:1024px; max-height:680px"',
             'salary_data': salary_data,
             'city_data': city_data,
             'header_year': header_year,
             'header_city': header_city,
             'profession_name': f"{ui.profession_name}",
             'h1_style': 'style="text-align:center; font-size:32px"',
             'h2_style': 'style="text-align:center"',
             'cell_style_none': "style=''",
             'cell_style': 'style="border:1px solid black; border-collapse: collapse; font-size: 16px; height: 19pt;'
                           'padding: 5px; text-align:center"'})

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, name, configuration=config, options={'enable-local-file-access': None})
    # endregion


def parse_html(line: str) -> str:
    line = sub('<.*?>', '', line)
    res = [' '.join(word.split()) for word in line.replace("\r\n", "\n").split('\n')]
    return res[0] if len(res) == 1 else res  # Спасибо Яндекс.Контесту за еще один костыль!


def parse_row_vacancy(row_vacs: list) -> dict:
    return dict(zip(title, map(parse_html, row_vacs)))


if __name__ == '__main__':
    translator = Translator()
    ui = UserInterface("vacancies_by_year.csv")
    csv = CSV(ui.file_name)
    title, row_vacancies = csv.title, csv.rows
    vacancies = [Vacancy(parse_row_vacancy(row_vac)) for row_vac in row_vacancies]
    ds = DataSet(vacancies, ui.profession_name)
    statistics = ds.get_data()
    report = Report(statistics)